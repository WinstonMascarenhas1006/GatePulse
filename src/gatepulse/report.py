"""Management exports. Sources: S-BI-01, S-JD-01 skill theme (decision documents) without copying JD domain."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from gatepulse.ai_risk import generate_insights
from gatepulse.config import DATA_EXPORTS, DATA_PROCESSED

HEADER_FILL = PatternFill("solid", fgColor="102A43")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _df_to_md(df: pd.DataFrame) -> str:
    try:
        return df.to_markdown(index=False)
    except ImportError:
        cols = list(df.columns)
        header = "| " + " | ".join(str(c) for c in cols) + " |"
        sep = "| " + " | ".join("---" for _ in cols) + " |"
        rows = ["| " + " | ".join(str(r[c]) for c in cols) + " |" for _, r in df.iterrows()]
        return "\n".join([header, sep, *rows])


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def export_excel(path: Path | None = None) -> Path:
    DATA_EXPORTS.mkdir(parents=True, exist_ok=True)
    path = path or DATA_EXPORTS / f"GatePulse_Steering_Pack_{datetime.now():%Y%m%d}.xlsx"

    facts = pd.read_csv(DATA_PROCESSED / "launch_facts_scored.csv")
    milestones = pd.read_csv(DATA_PROCESSED / "milestones.csv")
    issues = pd.read_csv(DATA_PROCESSED / "quality_issues.csv")
    risks = pd.read_csv(DATA_PROCESSED / "launch_risk_scores.csv")

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Executive_Summary"
    ws0["A1"] = "GatePulse - NPI Launch Readiness Steering Summary"
    ws0["A1"].font = Font(bold=True, size=14, color="102A43")
    ws0["A2"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    ws0["A4"] = "KPI"
    ws0["B4"] = "Value"
    _style_header(ws0)
    kpis = [
        ("Launches", len(facts)),
        ("Critical health", int((facts["health"] == "Critical").sum())),
        ("High AI slip risk", int((facts["slip_risk_label"] == "High").sum())),
        ("Avg progress %", round(float(facts["avg_progress"].mean()), 1)),
        ("Quality issues", len(issues)),
        ("Avg slip-risk score", round(float(facts["slip_risk_score"].mean()), 1)),
    ]
    for i, (k, v) in enumerate(kpis, start=5):
        ws0[f"A{i}"] = k
        ws0[f"B{i}"] = v

    insights = generate_insights(risks, facts)
    ws0["A12"] = "AI Insights"
    ws0["A12"].font = Font(bold=True)
    for i, text in enumerate(insights, start=13):
        ws0[f"A{i}"] = f"- {text}"
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    def add_df(name: str, df: pd.DataFrame) -> None:
        ws = wb.create_sheet(name[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        _style_header(ws)
        for col in ws.columns:
            maxlen = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 42)

    add_df("Launch_Facts", facts)
    add_df("Risk_Scores", risks.sort_values("slip_risk_score", ascending=False))
    add_df("Milestones", milestones)
    add_df("Quality_Issues", issues)
    wb.save(path)
    return path


def export_markdown_brief(path: Path | None = None) -> Path:
    DATA_EXPORTS.mkdir(parents=True, exist_ok=True)
    path = path or DATA_EXPORTS / "GatePulse_Executive_Brief.md"
    facts = pd.read_csv(DATA_PROCESSED / "launch_facts_scored.csv")
    risks = pd.read_csv(DATA_PROCESSED / "launch_risk_scores.csv")
    insights = generate_insights(risks, facts)
    top_risk = risks.sort_values("slip_risk_score", ascending=False).head(5)
    by_plant = (
        facts.groupby("plant_code")
        .agg(
            launches=("launch_id", "count"),
            avg_progress=("avg_progress", "mean"),
            avg_risk=("slip_risk_score", "mean"),
            critical=("health", lambda s: int((s == "Critical").sum())),
        )
        .round(1)
        .reset_index()
    )
    lines = [
        "# GatePulse Executive Brief - NPI Launch Portfolio",
        "",
        f"_Generated {datetime.now():%Y-%m-%d %H:%M}_",
        "",
        "## Purpose",
        "Decision-ready view of multi-plant launch gates, checklist data quality, and AI-assessed SOP slip risk.",
        "",
        "## Portfolio KPIs",
        f"- Launches in scope: **{len(facts)}**",
        f"- Critical health: **{int((facts['health']=='Critical').sum())}**",
        f"- High slip-risk (AI): **{int((facts['slip_risk_label']=='High').sum())}**",
        f"- Mean progress: **{facts['avg_progress'].mean():.1f}%**",
        "",
        "## Plant snapshot",
        "",
        _df_to_md(by_plant),
        "",
        "## Top 5 AI slip-risk launches",
        "",
        _df_to_md(
            top_risk[
                ["launch_name", "plant_code", "slip_risk_score", "slip_risk_label", "priority"]
            ]
        ),
        "",
        "## AI insights",
        "",
    ]
    for t in insights:
        lines.append(f"- {t}")
    lines += [
        "",
        "## Recommended next actions",
        "1. Unblock Critical launches before the next SOP gate review.",
        "2. Remediate open data-quality issues in checklists.",
        "3. Re-run GatePulse after the weekly plant sync.",
        "",
        "---",
        "_Synthetic demonstrator for Helion Industrial (fictional). Portfolio project._",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def export_all() -> dict[str, str]:
    return {"excel": str(export_excel()), "markdown": str(export_markdown_brief())}


if __name__ == "__main__":
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    print(export_all())
